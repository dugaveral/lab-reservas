import os
import psycopg2
import csv
from io import StringIO
from flask import Flask, render_template, request, redirect, Response, session
from datetime import datetime, timedelta
import random
import string
import sendgrid
from sendgrid.helpers.mail import Mail
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO


app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY") or "laboratorio-reservas-pades"
DB_URL = os.environ.get("DATABASE_URL")

EQUIPOS_LISTA = [
  "Agitador De Cabezal",
  "Agitador De Cabezal De Dispersión (Azul)",
  "Agitador De Cabezal De Dispersión (Azul) 2",
  "Agitador De Cabezal De Dispersión (Gris)",
  "Agitador De Cabezal De Dispersión (Gris) 2",
  "Agitador De Cabezal De Dispersión (Gris) 3",
  "Agitador Magnético Con Calentamiento",
  "Agitador Magnético",
  "Agitador Magnético 2",
  "Agitador Vortex",
  "Analizador De Tamaño De Partícula",
  "Analizador Elemental",
  "Autoclave",
  "Balanza 6Kg",
  "Balanza 6Kg 2",
  "Balanza De Humedad",
  "Balanza De Humedad 2",
  "Balanza De Humedad 3",
  "Balanza De Humedad 4",
  "Balanza Digital Capacidad 6 Kg Bacsa",
  "Balanza Gramera",
  "Balanzas Analítica",
  "Balanzas Analítica 2",
  "Balanzas Analítica 3",
  "Baño De Calentamiento Con Agitación",
  "Baño De María",
  "Baño Termostatado 12L 1",
  "Baño Termostatado 12L 2",
  "Baño Termostatado 8L",
  "Baño Termostatado Con Recirculación 24L",
  "Baño Ultrasónico",
  "Banda Trasportadora",
  "Batidora De 4.3 Litros",
  "Batidora De 4.3 Litros 2",
  "Batidora Marca",
  "Biorreactor De Lecho Fijo En Acero Inoxidable Tipo 304",
  "Bloque Digestor",
  "Bomba Calorimétrica",
  "Bomba De Vacío",
  "Bomba De Vacío 2",
  "Bomba De Vacío 3",
  "Bomba Isocratica / Cuaternaria",
  "Bomba Para Caldera",
  "Bomba Para Caldera 2",
  "Cámara De Estabilización",
  "Cámara De Fermentación",
  "Cámara De Fermentación 2",
  "Carretilla Para Afrecho En Acero Inoxidable Con Ruedas",
  "Centrifuga",
  "Centrifuga 2",
  "Centrifuga 3",
  "Shaker Con Agitación Orbital",
  "Colorímetro",
  "Colorímetro 2",
  "Compresor Elite",
  "Compresor Medic Air De 108 Litros",
  "Computador De Mesa Del Farinografo",
  "Congelador Horizontal",
  "Cromatógrafo De Gases Acoplado A Masas",
  "Cromatógrafo De Gases Con Detector FID",
  "Cúter",
  "Destilador De Agua",
  "Destilador De Agua 2",
  "Desengrasador 6 Puestos",
  "Deshumidificador",
  "Determinador De Actividad De Agua",
  "Determinador De Fibra",
  "Determinador De Fibra Tecnal",
  "Determinador De Grasa",
  "Determinador De Nitrógeno",
  "Embutidora Manual",
  "Empacadora Al Vacío",
  "Empacadora De Sólidos Semi Manual Con Sistema De Sellado Horizontal",
  "Empacadora Selladora",
  "Enfriador De Agua",
  "Equipo De Electrocoagulación",
  "Equipo De Ultrasonido",
  "Equipo Soxhlet De 6 Puestos",
  "Equipo Soxhlet De 6 Puestos 2",
  "Equipo Termoquímico Multipropósito",
  "Espectrofotómetro UV PerkinELMER",
  "Espectrofotómetro UV MAPADA",
  "Espectrofotómetro UV MERCK",
  "Estufa De Secado",
  "Extractor De Fibra",
  "Extrusora Marca Equipos Y Montajes",
  "Extrusora Para Pastas",
  "Farinografo",
  "Flexi Chill Set Line",
  "Floculador De 4 Puestos",
  "Floculador Portable",
  "Hidrociclón Acero Inoxidable Tipo 304 Con Bomba",
  "HPLC",	
  "Homogeneizador",
  "Homogeneizador 2",
  "Homogeneizador En Acero Inoxidable Aisi 304/Ha304-50150",
  "Horno 1",
  "Horno 2",
  "Horno 3",
  "Horno 4",
  "Horno 5",
  "Horno 6",
  "Horno 7",
  "Horno De Materia Volátil",
  "Incubadora Being",
  "Incubadora Being 2",
  "Licuadora",
  "Licuadora Osterizer",
  "Licuadora Osterizer 2",
  "Lavadora Peladora De Yuca",
  "Lavadora-Peladora De Vegetales",
  "Maquina Para Helado Suave Tres Boquillas Mh Twist 20 L",
  "Marmita Multipropósito Volcable Para Vapor",
  "Mouse Del HPLC",
  "Mezclador De Sólidos",
  "Micro Balanza",
  "Microscopio De Luz Polarizada",
  "Molino De Martillo Con Ciclón, Sistema De Carga Por Tornillo Alimentador Y Turbina",
  "Molino Multiusos",
  "Molino Multiusos 2",
  "Molino Para Carnes",
  "Molino Para Materiales Amiláceos",
  "Molino Tipo Willey",
  "Molino Tornillo Sin Fin",
  "Mufla 1",
  "Mufla 2",
  "Nevera",
  "Nevera 2",
  "Oxímetro",
  "Peletizadora (Negra)",
  "Peletizadora (Naranja)",
  "Ph-Metro",
  "Ph-Metro 2",
  "Ph-Metro 3",
  "Picador/Triturador",
  "Picadora De Materiales En Acero Inoxidable",
  "Pipeteador Recargable",
  "Pirolizador",
  "Plancha De Calentamiento Con Agitación",
  "Plancha De Calentamiento Con Agitación 2",
  "Plancha De Calentamiento Con Agitación 3",
  "Plancha De Calentamiento Con Agitación 4",
  "Plancha De Calentamiento Con Agitación 5",
  "Plancha De Calentamiento Con Agitación 6",
  "Plancha De Calentamiento Con Agitación 7",
  "Plancha De Calentamiento Con Agitación 8",
  "Plancha De Calentamiento Con Agitación 9",
  "Plancha De Calentamiento Con Agitación 10",
  "Plancha De Calentamiento Con Agitación 11",
  "Purificador De Agua",
  "Rallador De Yuca Con Motor",
  "Rack De Batería",
  "Reactor",
  "Reactor Con Agitación",
  "Reactor Multipropósito",
  "Refractómetro Digital",
  "Refractómetro Digital 2",
  "Regulador De Voltaje",
  "Regulador Eléctrico Con Elevador",
  "Reómetro Modular",
  "Sistema Canopy",
  "Sistema De Biogás Doméstico",
  "Sistema De Calentamiento De Agua Con Control Pid Y Bomba",
  "Sistema De Sensores Respirométricos 6 Mixi",
  "Sistema De Sensores Respirométricos 6 Mixi 2",
  "Sistema De Recuperación De Almidón Y Mucílagos De Ñame",
  "Sistema FTIR",
  "Tanque De Afrecho Cónico Construido En Acero Inoxidable",
  "Tanque De Alimentación Para Recirculación Y Almacenamiento De Almidón Con Sistema De Agitación",
  "Tanque De Mezclado De Soluciones De Glucosa En Acero Inoxidable",
  "Tanque Recolector De La Picadora En Acero Inoxidable",
  "Tanque Recolector De Lavadora En Acero Inoxidable",
  "Tanque Sedimentos En Tres Compartimentos En Acero Inoxidable",
  "Termo Reactor",
  "Texturómetro",
  "Unidad De DSC Set Line",
  "Unidad De Extrusión De Almidones Kecon",
  "Unidad De Fermentación - Módulo 1",
  "Unidad De Fermentación - Módulo 2",
  "Unidad De Secado Tipo Spray Vibrasec Sas",
  "Unidad Extrusión De Bioplásticos Bausano",
  "Unidad Neutralizadora Scrubber",
  "Zaranda",
  "Zaranda 2"
]

def generar_codigo():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def enviar_correo(destinatario, codigo):
    sg = sendgrid.SendGridAPIClient(api_key=os.environ.get('SENDGRID_API_KEY'))
    message = Mail(
        from_email='pades.reservas.laboratorios@gmail.com',
        to_emails=destinatario,
        subject='Código de tu reserva',
        html_content=f'<p>Tu código para eliminar la reserva es:</p><h2>{codigo}</h2>'
    )
    sg.send(message)

def get_db_connection():
    return psycopg2.connect(DB_URL)

def crear_columna_creado_en_si_no_existe():
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            ALTER TABLE reservas
            ADD COLUMN creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        """)
        conn.commit()
        print("✅ Columna 'creado_en' agregada exitosamente.")
    except psycopg2.errors.DuplicateColumn:
        print("ℹ️ La columna 'creado_en' ya existe.")
        conn.rollback()
    finally:
        cur.close()
        conn.close()

crear_columna_creado_en_si_no_existe()

@app.route('/')
def index():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT id, equipo, inicio, fin, usuario FROM reservas ORDER BY inicio')
    reservas = cur.fetchall()
    conn.close()
    return render_template('index.html', reservas=reservas)

@app.route('/reservar', methods=['GET', 'POST'])
def reservar():
    if request.method == 'POST':
        equipo = request.form['equipo']
        fecha = request.form['fecha']
        hora = request.form['hora']
        duracion_horas = int(request.form['duracion'])
        usuario = request.form['usuario']
        correo = request.form['correo']
        observaciones = request.form.get('observaciones', '')
        codigo = generar_codigo()

        inicio_dt = datetime.strptime(f"{fecha} {hora}", "%Y-%m-%d %H:%M")
        fin_dt = inicio_dt + timedelta(hours=duracion_horas)

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT 1 FROM reservas
            WHERE equipo = %s
            AND NOT (%s >= fin OR %s <= inicio)
        """, (equipo, inicio_dt, fin_dt))

        if cur.fetchone():
            conn.close()
            return render_template("error.html", mensaje="⚠️ Ya existe una reserva para ese equipo en ese periodo.")

        cur.execute("""
            INSERT INTO reservas (equipo, inicio, fin, usuario, codigo, observaciones)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (equipo, inicio_dt, fin_dt, usuario, codigo, observaciones))
        conn.commit()
        conn.close()

        try:
            enviar_correo(correo, codigo)
        except Exception as e:
            print("Error al enviar correo:", e)

        return render_template("codigo.html", codigo=codigo)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT equipo, inicio, fin, usuario FROM reservas ORDER BY equipo, inicio")
    registros = cur.fetchall()
    conn.close()

    reservas_por_equipo = {}
    for equipo, inicio, fin, usuario in registros:
        reservas_por_equipo.setdefault(equipo, []).append((inicio, fin, usuario))

    return render_template('reservar.html', equipos=EQUIPOS_LISTA, reservas_por_equipo=reservas_por_equipo)

@app.route('/reservas')
def ver_reservas():
    filtro_equipo = request.args.get('equipo')
    filtro_usuario = request.args.get('usuario')
    filtro_fecha = request.args.get('fecha')
    semana_inicio = request.args.get('semana_inicio')
    semana_fin = request.args.get('semana_fin')

    query = """
        SELECT id, equipo,
               TO_CHAR(inicio, 'YYYY-MM-DD HH24:MI'),
               TO_CHAR(fin, 'YYYY-MM-DD HH24:MI'),
               usuario,
               observaciones,
               TO_CHAR(creado_en, 'YYYY-MM-DD HH24:MI')
        FROM reservas
        WHERE TRUE
    """
    params = []

    if filtro_equipo:
        query += " AND equipo ILIKE %s"
        params.append(f"%{filtro_equipo}%")
    if filtro_usuario:
        query += " AND usuario ILIKE %s"
        params.append(f"%{filtro_usuario}%")
    if filtro_fecha:
        query += " AND DATE(inicio) = %s"
        params.append(filtro_fecha)
    if semana_inicio and semana_fin:
        try:
            inicio_dt = datetime.strptime(semana_inicio, "%Y-%m-%d")
            fin_dt = datetime.strptime(semana_fin, "%Y-%m-%d") + timedelta(days=1)
            query += " AND inicio >= %s AND inicio < %s"
            params.extend([inicio_dt, fin_dt])
        except ValueError:
            return render_template("error.html", mensaje="❌ Fechas inválidas.")

    query += " ORDER BY inicio"

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, params)
    reservas = cur.fetchall()
    conn.close()

    return render_template("ver_reservas.html", reservas=reservas, admin=session.get('admin'))

@app.route('/eliminar/<int:reserva_id>', methods=['POST'])
def eliminar_reserva(reserva_id):
    codigo = request.form['codigo'].strip().upper()

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT codigo FROM reservas WHERE id = %s', (reserva_id,))
    resultado = cur.fetchone()

    if resultado and resultado[0].strip().upper() == codigo:
        cur.execute('DELETE FROM reservas WHERE id = %s', (reserva_id,))
        conn.commit()
        mensaje = "✅ Reserva eliminada exitosamente."
    else:
        mensaje = "❌ Código incorrecto. No puedes eliminar esta reserva."

    conn.close()
    return render_template("error.html", mensaje=mensaje)

@app.route('/descargar', methods=['GET', 'POST'])
def descargar():
    if request.method == 'POST':
        password = request.form.get('password')
        if not password or password.strip() != "P4D3SADMIN#*":
            return render_template("error.html", mensaje="❌ Contraseña incorrecta para descargar.")

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, equipo, inicio, fin, usuario, observaciones, creado_en
            FROM reservas ORDER BY inicio
        """)
        rows = cur.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reservas"

        encabezados = ['ID', 'Equipo', 'Inicio', 'Fin', 'Responsable', 'Observaciones', 'Creado en']
        ws.append(encabezados)

        for fila in rows:
            ws.append(fila)

        for i, _ in enumerate(encabezados, 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": "attachment; filename=reservas.xlsx"}
        )

    return render_template("descargar.html")

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password.strip() == "P4D3SADMIN#*":
            session['admin'] = True
            return redirect('/admin')
        else:
            return render_template("error.html", mensaje="❌ Contraseña incorrecta.")
    return render_template("admin_login.html")

@app.route('/admin')
def admin():
    if not session.get('admin'):
        return redirect('/admin_login')

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, equipo,
               TO_CHAR(inicio, 'YYYY-MM-DD HH24:MI'),
               TO_CHAR(fin, 'YYYY-MM-DD HH24:MI'),
               usuario, observaciones,
               TO_CHAR(creado_en, 'YYYY-MM-DD HH24:MI')
        FROM reservas
        ORDER BY inicio
    """)
    reservas = cur.fetchall()
    conn.close()
    return render_template("ver_reservas.html", reservas=reservas, admin=True)

@app.route('/admin/eliminar/<int:reserva_id>', methods=['POST'])
def admin_eliminar_reserva(reserva_id):
    if not session.get('admin'):
        return redirect('/admin_login') 

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM reservas WHERE id = %s', (reserva_id,))
    conn.commit()
    conn.close()
    return redirect('/admin')

@app.route('/admin_logout')
def admin_logout():
    session.pop('admin', None)
    return redirect('/reservas')

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)